from __future__ import annotations

import csv
import io
import json
import re
from typing import Any

from utils.sqlite_schema import match_stage

MAX_QUESTIONS = 500
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

# 阶段 → 问题 id 前缀：决定流水线覆盖（05 推荐抽取 / 07 准确率）和三阶段归类
_STAGE_QID_TAG = {"symptom": "q4", "category": "q5", "brand": "q1"}

# 支持的列名映射（中英文均可）
COLUMN_ALIASES = {
    "question": "question",
    "问题": "question",
    "提问": "question",
    "geo关键词": "question",
    "关键词": "question",
    "product_name": "product",
    "product": "product",
    "产品": "product",
    "产品名": "product",
    "产品名称": "product",
    "level": "level",
    "层级": "level",
    "问题层级": "level",
    "scenario": "scenario",
    "场景": "scenario",
    "关键词类型": "scenario",
    "用户画像": "persona",
}


def _normalize_header(header: str) -> str | None:
    return COLUMN_ALIASES.get(str(header).strip().lower()) or COLUMN_ALIASES.get(str(header).strip())


def _slug(text: str, fallback: str) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Z]+", "", text)[:16]
    if cleaned:
        return cleaned.lower()
    if text.strip():
        # 纯中文产品名：取稳定 hash 作为 code，避免多个产品共用同一 code 互相覆盖
        import hashlib

        return "p" + hashlib.md5(text.strip().encode("utf-8")).hexdigest()[:8]
    return fallback


def _master_code_lookup() -> callable:
    """产品名 → 主数据 product_code。

    同一产品跨批次必须落在同一 code 上，趋势才连得起来；
    所以先按 products 主数据（名称+别名，双向包含）匹配，匹配不上才退回哈希 slug。
    """
    try:
        from . import product_master

        products = product_master.list_products(active_only=False)
    except Exception:
        products = []
    cache: dict[str, str | None] = {}

    def resolve(name: str) -> str | None:
        if not name:
            return None
        if name in cache:
            return cache[name]
        code = None
        for item in products:
            candidates = [item["product_name"], *item.get("aliases", [])]
            if any(c and (c in name or name in c) for c in candidates):
                code = item["product_code"]
                break
        cache[name] = code
        return code

    return resolve


def build_questions(
    rows: list[dict[str, str]], default_product: str = ""
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """行记录 → 标准问题列表 + 校验报告（层级分布/缺产品/未识别层级/去重）。"""
    questions: list[dict[str, Any]] = []
    seen: set[str] = set()
    last_product = ""
    master_code = _master_code_lookup()
    report = {
        "duplicates_removed": 0,
        "missing_product": 0,
        "missing_level": 0,
        "unknown_levels": set(),
        "stage_counts": {"symptom": 0, "category": 0, "brand": 0},
    }
    for row in rows:
        text = str(row.get("question") or "").strip()
        if not text:
            continue
        if text in seen:
            report["duplicates_removed"] += 1
            continue
        seen.add(text)
        # 产品列空时沿用上一行（Excel 合并单元格导出后只有首行有值）
        row_product = str(row.get("product") or "").strip()
        if row_product:
            last_product = row_product
        product = row_product or last_product or str(default_product or "").strip()
        if not product:
            report["missing_product"] += 1
        level = str(row.get("level") or "").strip()
        stage = match_stage(level)
        if not level:
            report["missing_level"] += 1
        elif stage is None:
            report["unknown_levels"].add(level)
        # 层级决定 id 前缀：病症→q4（05 推荐抽取）、品类→q5、品牌→q1（07 准确率+负面）
        stage = stage or "symptom"
        report["stage_counts"][stage] += 1
        index = len(questions) + 1
        product_code = (master_code(product) or _slug(product, "custom")) if product else "custom"
        questions.append({
            "id": f"{product_code}_{_STAGE_QID_TAG[stage]}_{index:04d}",
            "product": product,
            "product_code": product_code,
            "category": "user_upload",
            "level": level,
            "scenario": str(row.get("scenario") or "").strip(),
            "persona": str(row.get("persona") or "").strip(),
            "question": text,
            "has_brand_name": False,
            "is_variant": False,
            "variant_of": None,
        })
        if len(questions) > MAX_QUESTIONS:
            raise ValueError(f"问题数超过上限 {MAX_QUESTIONS}，请拆分后分批上传")
    if not questions:
        raise ValueError("未解析到任何问题，请检查文件格式（需包含“问题”或 question 列）")
    report["unknown_levels"] = sorted(report["unknown_levels"])
    return questions, report


def _rows_to_records(rows: list[list[str]]) -> list[dict[str, str]]:
    """在前 10 行内自动定位表头行（含“问题/GEO关键词”等列的行），其余行按表头映射。

    找不到表头时退化为“第一列即问题”。"""
    if not rows:
        return []
    header_index = -1
    headers: list[str | None] = []
    for i, row in enumerate(rows[:10]):
        candidate = [_normalize_header(h) for h in row]
        if "question" in candidate:
            header_index = i
            headers = candidate
            break
    if header_index < 0:
        return [{"question": row[0]} for row in rows if row and str(row[0]).strip()]
    result = []
    for row in rows[header_index + 1 :]:
        item: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if header and idx < len(row):
                item[header] = row[idx]
        result.append(item)
    return result


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return _rows_to_records(list(reader))


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [[("" if c is None else str(c)) for c in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return _rows_to_records(rows)


def _parse_json(content: bytes) -> list[dict[str, str]]:
    data = json.loads(content.decode("utf-8-sig", errors="replace"))
    if not isinstance(data, list):
        raise ValueError("JSON 必须是数组")
    rows = []
    for item in data:
        if isinstance(item, str):
            rows.append({"question": item})
        elif isinstance(item, dict):
            row: dict[str, str] = {}
            for key, value in item.items():
                normalized = _normalize_header(key)
                if normalized and value is not None:
                    row[normalized] = str(value)
            rows.append(row)
    return rows


def parse_upload(
    filename: str, content: bytes, default_product: str = ""
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """解析上传文件为标准问题列表 + 校验报告。"""
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError("文件超过 10MB")
    name = filename.lower()
    if name.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    elif name.endswith(".csv"):
        rows = _parse_csv(content)
    elif name.endswith(".json"):
        rows = _parse_json(content)
    else:
        raise ValueError("仅支持 .xlsx / .csv / .json 文件")
    return build_questions(rows, default_product)
